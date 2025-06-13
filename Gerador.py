################################################################################################################################################
# FERRAMENTA PARA DESENVOLVIMENTO DE SOFTWARES TIA PORTAL NO PADRÃO REDITECH
# DATA DE INICIO: 25/04/2024
# Autor: Caio Ribeiro
# Colaboradores: Willie, Enzo Oliveira e Gustavo Lourenção
# Versão 0 - 25/04/2024 - Versão Inicial
# Versão 1 - 26/04/2024 - Alterado para ler arquivos excel
# Versão 2 - 27/04/2024 - Adicionado diretório relativo ao lugar salvo e geração de arquivos de chamada de função, comentarios adicionados
# Versão 3 - 28/04/2024 - Adicionado geração de alarmes IHM
# Versão 4 - 29/04/2024 - Adicionado geração de interface com IOs
# Versão 5 - 08/08/2024 - Adicionado interface grafica para o usuario GUI
# Versão 6 - 20/09/2024 - Corrigido os textos das DBS da planilha "Alarmes IHM" - Retirado o campo "Alarm text [en=US]" da planilha, pois estava dando conflito com o TIA 18
# Versão 7 - 23/01/2025 - Corrigido a lógica para criação das DBs do telegrama 20 e adicionado tags padrão
# Versão 8 - 22/05/2025 - Adicionado a lógica de criação de tabelas de PopUps HMI e HMI Objects
# Versão 9 - 28/05/2025 - Adicionado a lógica de criação de tabelas do Index HMI
# Versão 10 - 02/06/2025 - ~Adicionado lógica para novo tipo de alarme - LLL e HHH - para o projeto azulão~ - #####Não aplicável nessa versão######
# Versão 11 - 03/06/2026 - Adicionado lógica para a criação dos textos dos itens da IHM
################################################################################################################################################


# Importando bibliotecas
import openpyxl # Biblioteca para abrir/ler/editar arquivos excel
import os       # Biblioteca para conseguir usar endereçamento de diretórios relativamente a pasta aonde o programa esta salvo 
import sys
import shutil
from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QFileDialog
from PyQt5.QtCore import QFile
from PyQt5.QtWidgets import QLabel

# Atribuindo o valor do diretório atual na variavel current_directory
current_directory = os.path.realpath(os.path.join(os.getcwd(), os.path.dirname(__file__)))
print("Diretório relativo: ", current_directory)
#current_directory = os.getcwd()

# Diretório dos arquivos de programa montados com relativismo ao diretório dos arquivos de referência
#file_path_mapa_de_memoria = os.path.join(current_directory,"GERADOR\ELEMENTS\LISTA DE ELEMENTOS.xlsx")
file_path_FC_CALL = os.path.join(current_directory, "SOURCE\\Call_Funtions.TXT")
file_path_FC_DXV = os.path.join(current_directory, "SOURCE\\FC_DXV.TXT")
file_path_FC_AV = os.path.join(current_directory, "SOURCE\\FC_AV.TXT")
file_path_FC_ACM = os.path.join(current_directory, "SOURCE\\FC_ACM.TXT")
file_path_FC_AXV = os.path.join(current_directory, "SOURCE\\FC_AXV.TXT")
file_path_FC_DCM = os.path.join(current_directory, "SOURCE\\FC_DCM.TXT")
file_path_FC_DI = os.path.join(current_directory, "SOURCE\\FC_DI.TXT")
file_path_FC_AREA = os.path.join(current_directory, "SOURCE\\FC_AREA.TXT")
file_path_DB_DXV = os.path.join(current_directory, "SOURCE\\DB_DXV.db")
file_path_DB_AV = os.path.join(current_directory, "SOURCE\\DB_AV.db")
file_path_DB_ACM = os.path.join(current_directory, "SOURCE\\DB_ACM.db")
file_path_DB_G120C_Tel20 = os.path.join(current_directory, "SOURCE\\DB_G120C_Tel20.db")
file_path_DB_DCM = os.path.join(current_directory, "SOURCE\\DB_DCM.db")
file_path_DB_AXV = os.path.join(current_directory, "SOURCE\\DB_AXV.db")
file_path_DB_DI = os.path.join(current_directory, "SOURCE\\DB_DI.db")
file_path_DB_AREA = os.path.join(current_directory, "SOURCE\\DB_AREA.db")
file_path_FBS = os.path.join(current_directory, "SOURCE\\FBs\\fbs.scl")
print("Diretorio file_path_FBS: " ,file_path_FBS)
# Definição de strings de controle das colunas dentro do arquivo excel
tag_name = "CODIGO KKS (TAG)"
função_name = "FUNCAO"
fc_db_number = "DB"
tipo_controle = "BLOCO DE CONTROLE"
area_controle = "AREA"
ENDEREÇO_number = "ENDEREÇO"
tipo_controle_tag = "TIPO"

# Função para LER o arquivo excel de acordo com as strings de controle de coluna
def read_column(file_path, column_name, columm_name_check):
    wb = openpyxl.load_workbook(file_path)
    
    sheet = wb[columm_name_check]
    column_index = {}

    for cell in sheet[1]:
        column_index[cell.value] = cell.column

    if column_name not in column_index:
        print(f"Column '{column_name}' not found.")
        return []
    
    column_values = []
    column_letter = openpyxl.utils.get_column_letter(column_index[column_name])
    for cell in sheet[column_letter]:
            if (cell.value) != column_name:
                column_values.append(cell.value)
    return column_values

# Função para gerar lógica apartir de modelos e relacionar com os dados levantado das colunas de excel controloadas pelas strings de controle
def generate_logic(pathmodel, tag_data, area_data, função_data, fc_db_number_data, tipo_controle_data, type, block):
    FC_fct_array = []
    FC_fct_array_t = []
    FC_REPLACED_array = []    
    FC_REPLACED_AREA_array = []
    FC_REPLACED_TITULO_array = []
    FC_REPLACED_NUMERO_array = []
    FC_REPLACED_COMENTARIO_array = []
    with open(pathmodel, "r") as file:
        for line in file:
            FC_fct_array.append(line.strip()) 
    for j in range(len(tag_data)):
        FC_fct_array_t = FC_fct_array
        if(tipo_controle_data[j] == type):
            for i in range(len(FC_fct_array)):
                FC_REPLACED_array.append(FC_fct_array_t[i].replace("TAG", str(fc_db_number_data[j])))
                FC_REPLACED_AREA_array.append(FC_REPLACED_array[i].replace("AREA", str(area_data[j])))
                FC_REPLACED_TITULO_array.append(FC_REPLACED_AREA_array[i].replace("Titulo", str(função_data[j])))
                FC_REPLACED_NUMERO_array.append(FC_REPLACED_TITULO_array[i].replace("POSICAO", str(fc_db_number_data[j])))
                FC_REPLACED_COMENTARIO_array.append(FC_REPLACED_NUMERO_array[i].replace("COMENTARIO", str(função_data[j])))

            if block == "DB":            
                    file_path = os.path.join(current_directory, "GENERATED/DB_"+tag_data[j]+".DB")
                    print(file_path)
                    with open(file_path, "w") as file:
                        for line in FC_REPLACED_COMENTARIO_array:
                            file.write(line + "\n")
                    FC_fct_array_t =[]
                    FC_REPLACED_array = []
                    FC_REPLACED_AREA_array = []
                    FC_REPLACED_TITULO_array = []
                    FC_REPLACED_NUMERO_array = []
                    FC_REPLACED_COMENTARIO_array = []
            if block == "FC":            
                    file_path = os.path.join(current_directory, "GENERATED/FC_"+tag_data[j]+".SCL")
                    print(file_path)
                    with open(file_path, "w") as file:
                        for line in FC_REPLACED_COMENTARIO_array:
                            file.write(line + "\n")
                    FC_fct_array_t =[]
                    FC_REPLACED_array = []
                    FC_REPLACED_AREA_array = []
                    FC_REPLACED_TITULO_array = []
                    FC_REPLACED_NUMERO_array = []
                    FC_REPLACED_COMENTARIO_array = []

def generate_logic_acm_telegram_g20(pathmodel, tag_data, area_data, função_data, fc_db_number_data, tipo_controle_data, type, block):
    FC_fct_array = []
    FC_fct_array_t = []
    FC_REPLACED_array = []    
    FC_REPLACED_AREA_array = []
    FC_REPLACED_TITULO_array = []
    FC_REPLACED_NUMERO_array = []
    FC_REPLACED_COMENTARIO_array = []
    FC_REPLACED_TAGDATA_array = []
    with open(pathmodel, "r") as file:
        for line in file:
            FC_fct_array.append(line.strip()) 
    for j in range(len(tag_data)):
        FC_fct_array_t = FC_fct_array
        if(tipo_controle_data[j] == type):
            for i in range(len(FC_fct_array)):
                FC_REPLACED_array.append(FC_fct_array_t[i].replace("TAG", str(fc_db_number_data[j])))
                FC_REPLACED_AREA_array.append(FC_REPLACED_array[i].replace("AREA", str(area_data[j])))
                FC_REPLACED_TITULO_array.append(FC_REPLACED_AREA_array[i].replace("Titulo", str(função_data[j])))
                FC_REPLACED_NUMERO_array.append(FC_REPLACED_TITULO_array[i].replace("POSICAO", str(fc_db_number_data[j])))
                FC_REPLACED_COMENTARIO_array.append(FC_REPLACED_NUMERO_array[i].replace("COMENTARIO", str(função_data[j])))
                FC_REPLACED_TAGDATA_array.append(FC_REPLACED_COMENTARIO_array[i].replace("DATA", str(tag_data[j])))
            if block == "DB":            
                    file_path = os.path.join(current_directory, "GENERATED/DB_"+tag_data[j]+".DB")
                    print(file_path)
                    with open(file_path, "w") as file:
                        for line in FC_REPLACED_COMENTARIO_array:
                            file.write(line + "\n")
                    FC_fct_array_t =[]
                    FC_REPLACED_array = []
                    FC_REPLACED_AREA_array = []
                    FC_REPLACED_TITULO_array = []
                    FC_REPLACED_NUMERO_array = []
                    FC_REPLACED_COMENTARIO_array = []
            if block == "FC":            
                    file_path = os.path.join(current_directory, "GENERATED/FC_"+tag_data[j]+".SCL")
                    print(file_path)
                    with open(file_path, "w") as file:
                        for line in FC_REPLACED_TAGDATA_array:
                            file.write(line + "\n")
                    FC_fct_array_t =[]
                    FC_REPLACED_array = []
                    FC_REPLACED_AREA_array = []
                    FC_REPLACED_TITULO_array = []
                    FC_REPLACED_NUMERO_array = []
                    FC_REPLACED_COMENTARIO_array = []
                    FC_REPLACED_TAGDATA_array = []


# Função para gerar lóciga de chamada de função e relacionar com os dados levantado das colunas de excel controloadas pelas strings de controle 
def generate_call(pathmodel, tag_data, função_data, tipo_controle_data, fc_db_number_data):
    FC_fct_array = []
    FC_fct_replace_array = []
    FC_fct_replace_FUNCOES_array =[]
    with open(pathmodel, "r") as file:
        for line in file:
            if line.strip() == "####":
                for i in range(len(tag_data)):
                    if(função_data[i]==tipo_controle_data):
                        FC_fct_array.append("FC"+str(fc_db_number_data[i])+"();")
            else :
                FC_fct_array.append(line)
    for i in range(len(FC_fct_array)):
        FC_fct_replace_array.append(FC_fct_array[i].replace("\n",''))
        FC_fct_replace_FUNCOES_array.append(FC_fct_replace_array[i].replace("FUNCOES","FC_CALL_"+tipo_controle_data))
    
    file_path = os.path.join(current_directory, "GENERATED/FC_CALL_"+tipo_controle_data+".SCL")

    with open(file_path, "w") as file:
        for line in FC_fct_replace_FUNCOES_array:
            file.write(line + "\n")
    print(file_path)

class MainWin(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setGeometry(200, 200, 300, 380)
        self.setWindowTitle("Gerador de códigos")
                
        self.status_label = QLabel("Status: Selecione um arquivo!", self)
        self.status_label.resize(200, 32)
        self.status_label.move(10, 10)

        button = QPushButton('Abrir arquivo', self)
        button.clicked.connect(self.open_file)
        button.resize(120, 32)
        button.move(100, 50)

        button1 = QPushButton('Gerar Arquivos', self)
        button1.clicked.connect(self.Generate)
        button1.resize(120, 32)
        button1.move(100, 90)

        button2 = QPushButton('Gerar Alarmes', self)
        button2.clicked.connect(self.Alarmes)
        button2.resize(120, 32)
        button2.move(100, 130)
        
        button3 = QPushButton('Gerar Tags PLC', self)
        button3.clicked.connect(self.Tags)
        button3.resize(120, 32)
        button3.move(100, 170)

        button4 = QPushButton('Gerar Tags HMI', self)
        button4.clicked.connect(self.HMI_Tags)
        button4.resize(120, 32)
        button4.move(100, 210)

        button5 = QPushButton('Gerar Tags Objects', self)
        button5.clicked.connect(self.Tags_Objects)
        button5.resize(120, 32)
        button5.move(100, 250)

        button6 = QPushButton('Gerar Tags Index', self)
        button6.clicked.connect(self.Tags_Index)
        button6.resize(120, 32)
        button6.move(100, 290)

        button7 = QPushButton('Gerar Textos IHM', self)
        button7.clicked.connect(self.Textos_IHM)
        button7.resize(120,32)
        button7.move(100,330)

        self.show()
        self.filename = ''

    def open_file(self):
        self.filename, _ = QFileDialog.getOpenFileName(self, "Abrir arquivo excel", "", "Extensão excel (*.xlsx)")
        if self.filename != '':
            self.status_label.setText("Status: Arquivo aberto")

    def Generate(self):
        
        if self.filename != '':
            File_array = []
            with open(file_path_FBS, "r") as file:
                for line in file:
                    File_array.append(line.strip())
                file_path = os.path.join(current_directory, "GENERATED/FBS.SCL")
            with open(file_path, "w") as file:
                for line in File_array:
                    file.write(line + "\n")    
            print(file_path)

            # Chamada da função de leitura de do arquivo de excel
            tag_data = read_column(self.filename, tag_name, 'LISTA DE MEMORIA')
            função_data = read_column(self.filename, função_name, 'LISTA DE MEMORIA')
            tipo_controle_data = read_column(self.filename, tipo_controle, 'LISTA DE MEMORIA')
            fc_db_number_data = read_column(self.filename, fc_db_number, 'LISTA DE MEMORIA')
            area_data = read_column(self.filename, area_controle, 'LISTA DE MEMORIA')

            # Chamada de função para gerar chamada e lógica
            generate_call(file_path_FC_CALL, tag_data, tipo_controle_data, "DXV", fc_db_number_data)
            generate_call(file_path_FC_CALL, tag_data, tipo_controle_data, "ACM", fc_db_number_data)
            generate_call(file_path_FC_CALL, tag_data, tipo_controle_data, "DCM", fc_db_number_data)
            generate_call(file_path_FC_CALL, tag_data, tipo_controle_data, "AV", fc_db_number_data)
            generate_call(file_path_FC_CALL, tag_data, tipo_controle_data, "DI", fc_db_number_data)
            generate_call(file_path_FC_CALL, tag_data, tipo_controle_data, "AXV", fc_db_number_data)

            # Chamada de função para gerar lógica de area
            generate_logic(file_path_FC_AREA, tag_data, area_data, função_data, fc_db_number_data, tipo_controle_data, "AREA", "FC")
            generate_logic(file_path_DB_AREA, tag_data, area_data, função_data, fc_db_number_data, tipo_controle_data, "AREA", "DB")

            # Chamada da função de gerar lógica para cada item existente do padrão
            generate_logic(file_path_FC_DXV, tag_data, area_data, função_data, fc_db_number_data, tipo_controle_data, "DXV", "FC")
            generate_logic(file_path_DB_DXV, tag_data, area_data, função_data, fc_db_number_data, tipo_controle_data, "DXV", "DB")
            generate_logic(file_path_FC_AV, tag_data, area_data, função_data, fc_db_number_data, tipo_controle_data, "AV", "FC")
            generate_logic(file_path_DB_AV, tag_data, area_data, função_data, fc_db_number_data, tipo_controle_data, "AV", "DB")
            generate_logic_acm_telegram_g20(file_path_FC_ACM, tag_data, area_data, função_data, fc_db_number_data, tipo_controle_data, "ACM", "FC")
            generate_logic(file_path_DB_ACM, tag_data, area_data, função_data, fc_db_number_data, tipo_controle_data, "ACM", "DB")
            generate_logic(file_path_DB_G120C_Tel20, tag_data, area_data, função_data, fc_db_number_data, tipo_controle_data, "T20", "DB")
            generate_logic(file_path_FC_DCM, tag_data, area_data, função_data, fc_db_number_data, tipo_controle_data, "DCM", "FC")
            generate_logic(file_path_DB_DCM, tag_data, area_data, função_data, fc_db_number_data, tipo_controle_data, "DCM", "DB")
            generate_logic(file_path_FC_DI, tag_data, area_data, função_data, fc_db_number_data, tipo_controle_data, "DI", "FC")
            generate_logic(file_path_DB_DI, tag_data, area_data, função_data, fc_db_number_data, tipo_controle_data, "DI", "DB")
            generate_logic(file_path_FC_AXV, tag_data, area_data, função_data, fc_db_number_data, tipo_controle_data, "AXV", "FC")
            generate_logic(file_path_DB_AXV, tag_data, area_data, função_data, fc_db_number_data, tipo_controle_data, "AXV", "DB")

            print(self.filename)
            print("Arquivos gerados com SUCESSO!!!")
            self.status_label.setText("Status: Arquivos gerados")

    def Tags_Index(self):
        if self.filename != '':
            # Define a tabela aberta ativa
            workbookObjects = openpyxl.Workbook()
            sheetObjects = workbookObjects.active
            sheetObjects.title = 'Hmi Tags'
            tag_data = read_column(self.filename, tag_name, 'LISTA DE MEMORIA')
            função_data = read_column(self.filename, função_name, 'LISTA DE MEMORIA')
            tipo_controle_data = read_column(self.filename, tipo_controle, 'LISTA DE MEMORIA')
            fc_db_number_data = read_column(self.filename, fc_db_number, 'LISTA DE MEMORIA')
            area_data = read_column(self.filename, area_controle, 'LISTA DE MEMORIA')
            tag_data = read_column(self.filename, tag_name, 'LISTA DE MEMORIA')
            adicionados_fp_screen_pointer = set()   
            cabecalho = ["Name", "Path", "Connection", "PLC tag", "DataType", "Length", "Coding", "Access Method", "Address", "Indirect addressing", "Index tag", "Start value", "ID tag", "Display name [en-US]", "Comment [en-US]", "Acquisition mode", "Acquisition cycle", "Limit Upper 2 Type", "Limit Upper 2", "Limit Upper 1 Type", "Limit Upper 1", "Limit Lower 1 Type", "Limit Lower 1", "Limit Lower 2 Type","Limit Lower 2", "Linear scaling", "End value PLC", "Start value PLC", "End value HMI", "Start value HMI", "Gmp relevant", "Confirmation Type", "Mandatory Commenting"]
            sheetObjects.append(cabecalho)
            # Write data to the worksheet
            tipos_permitidos = {"AV", "ACM", "AXV", "DI", "DCM", "DXV"}
        for i in range(len(tag_data)):
            tag = str(tag_data[i]) if tag_data[i] else "TAG"
            tipo = str(tipo_controle_data[i]) if tipo_controle_data[i] else "TIPO"
            if tipo not in tipos_permitidos:
                continue
            fc_db = str(fc_db_number_data[i]) if fc_db_number_data[i] else "0"
            path = f"{tipo}\\{tipo}_Index"

            # Sempre adiciona a linha de index
            sheetObjects.append([
                f"{tipo}_Index_" + tag, path, "<No Value>", "<No Value>", "Int", "2", "Binary",
                "<No Value>", "<No value>", "False", "<No value>", str(fc_db_number_data[i]), "0", "<No value>", "<No value>",
                "Cyclic in operation", "1s", "None", "<No value>", "None", "<No value>", "None", "<No value>", "None",
                "<No value>", "False", "10", "0", "100", "0", "False", "None", "False"
            ])

                # Adiciona FP_Screen_Pointer somente uma vez por tipo
            if tipo not in adicionados_fp_screen_pointer:
             sheetObjects.append([
                f"{tipo}_FP_Screen_Pointer", path, "<No Value>", "<No Value>", "Int", "2", "Binary",
                "<No Value>", "<No value>", "False", "<No value>", "None", "0", "<No value>", "<No value>",
                "Cyclic in operation", "1s", "None", "<No value>", "None", "<No value>", "None", "<No value>", "None",
                "<No value>", "False", "10", "0", "100", "0", "False", "None", "False"
            ])
            adicionados_fp_screen_pointer.add(tipo)
        PATH = os.path.join(current_directory,"GENERATED\\HMI_Index.xlsx")
        workbookObjects.save(PATH)
        workbookObjects.close()
        self.status_label.setText("Status: Lista 'Index' Gerada")
        print(PATH)
        print("O arquivo de Excel 'HMI_Index.xlsx' foi criado com sucesso!!")

    def Textos_IHM(self):
        if self.filename != '':
            workbook = openpyxl.Workbook()
            # Cria a primeira aba: TextListEntry
            sheetTexts = workbook.active
            sheetTexts.title = 'TextList'
            tag_data = read_column(self.filename, tag_name, 'LISTA DE MEMORIA')
            função_data = read_column(self.filename, função_name, 'LISTA DE MEMORIA')
            tipo_controle_data = read_column(self.filename, tipo_controle, 'LISTA DE MEMORIA')
            fc_db_number_data = read_column(self.filename, fc_db_number, 'LISTA DE MEMORIA')
            area_data = read_column(self.filename, area_controle, 'LISTA DE MEMORIA')
            cabecalho2 = ["Name", "ListRange", "Comment [en-US]"]
            sheetTexts.append(cabecalho2)
            sheetTextlist = workbook.create_sheet(title='TextListEntry')
            cabecalho = ["Name", "Parent", "DefaultEntry", "Value", "Text [en-US]", "FieldInfos"]
            sheetTextlist.append(cabecalho)
            tipos_adicionados = set()
            for tipo in tipo_controle_data:
                if tipo in tipos_adicionados:
                    continue  # pula se já adicionou esse tipo
                if tipo == "ACM":
                    sheetTexts.append(["CM_ACM_Tagname","Decimal", "<No value>"])
                    sheetTexts.append(["CM_ACM_Description","Decimal", "<No value>"])
                    sheetTexts.append(["CM_ACM_Button_Pos1_Text","Decimal", "<No value>"])
                    sheetTexts.append(["CM_ACM_Button_Stop_Text","Decimal", "<No value>"])
                    sheetTexts.append(["CM_ACM_Button_Pos0_Text","Decimal", "<No value>"])
                elif tipo == "DCM":
                    sheetTexts.append(["CM_DCM_Tagname","Decimal", "<No value>"])
                    sheetTexts.append(["CM_DCM_Description","Decimal", "<No value>"])
                    sheetTexts.append(["CM_DCM_Button_Pos1_Text","Decimal", "<No value>"])
                    sheetTexts.append(["CM_DCM_Button_Stop_Text","Decimal", "<No value>"])
                    sheetTexts.append(["CM_DCM_Button_Pos0_Text","Decimal", "<No value>"])
                elif tipo == "DXV":
                    sheetTexts.append(["CM_DXV_Tagname","Decimal", "<No value>"])
                    sheetTexts.append(["CM_DXV_Description","Decimal", "<No value>"])
                    sheetTexts.append(["CM_DXV_Button_Pos1_Text","Decimal", "<No value>"])
                    sheetTexts.append(["CM_DXV_Button_Pos0_Text","Decimal", "<No value>"])
                elif tipo == "DI":
                    sheetTexts.append(["CM_DI_Tagname","Decimal", "<No value>"])
                    sheetTexts.append(["CM_DI_Description","Decimal", "<No value>"])
                    sheetTexts.append(["CM_DI_Device_Type_Text","Decimal", "<No value>"])
                elif tipo == "AXV":
                    sheetTexts.append(["CM_AXV_Tagname","Decimal", "<No value>"])
                    sheetTexts.append(["CM_AXV_Description","Decimal", "<No value>"])
                elif tipo == "AV":
                    sheetTexts.append(["CM_AV_Tagname","Decimal", "<No value>"])
                    sheetTexts.append(["CM_AV_Description","Decimal", "<No value>"])
                    sheetTexts.append(["CM_AV_Device_Type_Text","Decimal", "<No value>"])
                    sheetTexts.append(["CM_AV_Dimension_Text","Decimal", "<No value>"])
                    sheetTexts.append(["CM_AV_Dimension_Totalizer","Decimal", "<No value>"])
                tipos_adicionados.add(tipo)  # marca como processado
            PATH = os.path.join(current_directory,"GENERATED\\Textos_IHM_Passo1.xlsx")
            workbook.save(PATH)
            workbook.close()
            self.status_label.setText("Status: Lista de textos Gerada")
            print(PATH)
            print("O arquivo de Excel 'Textos_IHM_Passo1.xlsx' foi criado com sucesso!!")
                # Cria a segunda aba: TextList
            workbook = openpyxl.Workbook()
            sheetTexts = workbook.active
            sheetTexts.title = 'TextList'
            cabecalho2 = ["Name", "ListRange", "Comment [en-US]"]
            sheetTexts.append(cabecalho2)
            tag_data = read_column(self.filename, tag_name, 'LISTA DE MEMORIA')
            função_data = read_column(self.filename, função_name, 'LISTA DE MEMORIA')
            tipo_controle_data = read_column(self.filename, tipo_controle, 'LISTA DE MEMORIA')
            fc_db_number_data = read_column(self.filename, fc_db_number, 'LISTA DE MEMORIA')
            area_data = read_column(self.filename, area_controle, 'LISTA DE MEMORIA')
            sheetTextlist = workbook.create_sheet(title='TextListEntry')
            cabecalho = ["Name", "Parent", "DefaultEntry", "Value", "Text [en-US]", "FieldInfos"]
            sheetTextlist.append(cabecalho)
            contador_ACM_Tagname = 1
            contador_ACM_Description = 1
            contador_ACM_pos1 = 1
            contador_ACM_pos0 = 1
            contador_ACM_stop = 1
            contador_DCM_Tagname = 1
            contador_DCM_Description = 1
            contador_DCM_pos1 = 1
            contador_DCM_pos0 = 1
            contador_DCM_stop = 1
            contador_AV_Tagname = 1
            contador_AV_Description = 1
            contador_AV_dimension_totalizer = 1
            contador_AV_dimension = 1
            contador_AV_device_type = 1
            contador_DI_Description = 1
            contador_DI_device_type = 1
            contador_DI_Tagname = 1
            contador_DXV_Tagname = 1
            contador_DXV_Description = 1
            contador_DXV_pos1 = 1
            contador_DXV_pos0 = 1
            contador_AXV_Tagname = 1
            contador_AXV_Description = 1
            # Write data to the worksheet
            for j in range(len(tag_data)):
                if(tipo_controle_data[j] == "ACM"):
                     sheetTextlist.append([f"Text_list_entry_{contador_ACM_Tagname}", "CM_ACM_Tagname", "", fc_db_number_data[j], tag_data[j],""])
                     contador_ACM_Tagname += 1
                     sheetTextlist.append([f"Text_list_entry_{contador_ACM_Description}", "CM_ACM_Description", "", fc_db_number_data[j], função_data[j],""])
                     contador_ACM_Description += 1
                     sheetTextlist.append([f"Text_list_entry_{contador_ACM_pos1}", "CM_ACM_Button_Pos1_Text", "", fc_db_number_data[j], "REV", ""])
                     contador_ACM_pos1 += 1
                     sheetTextlist.append([f"Text_list_entry_{contador_ACM_stop}", "CM_ACM_Button_Stop_Text","",fc_db_number_data[j], "Desl.", ""])
                     contador_ACM_stop += 1
                     sheetTextlist.append([f"Text_list_entry_{contador_ACM_pos0}", "CM_ACM_Button_Pos0_Text","",fc_db_number_data[j], "Ligar", ""])
                     contador_ACM_pos0 += 1
                if(tipo_controle_data[j] == "DCM"):
                     sheetTextlist.append([f"Text_list_entry_{contador_DCM_Tagname}", "CM_DCM_Tagname", "", fc_db_number_data[j], tag_data[j],""])
                     contador_DCM_Tagname += 1
                     sheetTextlist.append([f"Text_list_entry_{contador_DCM_Description}", "CM_DCM_Description", "", fc_db_number_data[j], função_data[j],""])
                     contador_DCM_Description += 1
                     sheetTextlist.append([f"Text_list_entry_{contador_DCM_pos1}", "CM_DCM_Button_Pos1_Text", "", fc_db_number_data[j], "REV", ""])
                     contador_DCM_pos1 += 1
                     sheetTextlist.append([f"Text_list_entry_{contador_DCM_stop}", "CM_DCM_Button_Stop_Text","",fc_db_number_data[j], "Desl.", ""])
                     contador_DCM_stop += 1
                     sheetTextlist.append([f"Text_list_entry_{contador_DCM_pos0}", "CM_DCM_Button_Pos0_Text","",fc_db_number_data[j], "Ligar", ""])
                     contador_DCM_pos0 += 1
                if(tipo_controle_data[j] == "AV"):
                     sheetTextlist.append([f"Text_list_entry_{contador_AV_Tagname}", "CM_AV_Tagname", "", fc_db_number_data[j], tag_data[j],""])
                     contador_AV_Tagname += 1
                     sheetTextlist.append([f"Text_list_entry_{contador_AV_Description}", "CM_AV_Description", "", fc_db_number_data[j], função_data[j],""])
                     contador_AV_Description += 1
                     sheetTextlist.append([f"Text_list_entry_{contador_AV_dimension_totalizer}", "CM_AV_Dimension_Totalizer_Text", "", fc_db_number_data[j], "%", ""])
                     contador_AV_dimension_totalizer += 1
                     sheetTextlist.append([f"Text_list_entry_{contador_AV_device_type}", "CM_AV_Device_Type_Text","",fc_db_number_data[j], "FT", ""])
                     contador_AV_device_type += 1
                     sheetTextlist.append([f"Text_list_entry_{contador_AV_dimension}", "CM_AV_Dimension_Text","",fc_db_number_data[j], "m³/h", ""])
                     contador_AV_dimension += 1
                if(tipo_controle_data[j] == "DI"):
                     sheetTextlist.append([f"Text_list_entry_{contador_DI_Tagname}", "CM_DI_Tagname", "", fc_db_number_data[j], tag_data[j],""])
                     contador_DI_Tagname += 1
                     sheetTextlist.append([f"Text_list_entry_{contador_DI_Description}", "CM_DI_Description", "", fc_db_number_data[j], função_data[j],""])
                     contador_DI_Description += 1
                     sheetTextlist.append([f"Text_list_entry_{contador_DI_device_type}", "CM_DI_Device_Type_Text","",fc_db_number_data[j], "FS", ""])
                     contador_DI_device_type += 1
                if(tipo_controle_data[j] == "DXV"):
                     sheetTextlist.append([f"Text_list_entry_{contador_DXV_Tagname}", "CM_DXV_Tagname", "", fc_db_number_data[j], tag_data[j],""])
                     contador_DXV_Tagname += 1
                     sheetTextlist.append([f"Text_list_entry_{contador_DXV_Description}", "CM_DXV_Description", "", fc_db_number_data[j], função_data[j],""])
                     contador_DXV_Description += 1
                     sheetTextlist.append([f"Text_list_entry_{contador_DXV_pos1}", "CM_DXV_Button_Pos1_Text", "", fc_db_number_data[j], "Abrir", ""])
                     contador_DXV_pos1 += 1
                     sheetTextlist.append([f"Text_list_entry_{contador_DXV_pos0}", "CM_DXV_Button_Pos0_Text","",fc_db_number_data[j], "Fechar", ""])
                     contador_DXV_pos0 += 1
                if(tipo_controle_data[j] == "AXV"):
                     sheetTextlist.append([f"Text_list_entry_{contador_AXV_Tagname}", "CM_AXV_Tagname", "", fc_db_number_data[j], tag_data[j],""])
                     contador_AXV_Tagname += 1
                     sheetTextlist.append([f"Text_list_entry_{contador_AXV_Description}", "CM_AXV_Description", "", fc_db_number_data[j], função_data[j],""])
                     contador_AXV_Description += 1
            PATH = os.path.join(current_directory,"GENERATED\\Textos_IHM_Passo2.xlsx")
            workbook.save(PATH)
            workbook.close()
            self.status_label.setText("Status: Lista de textos Gerada")
            print(PATH)
            print("O arquivo de Excel 'Textos_IHM_Passo2.xlsx' foi criado com sucesso!!")

    def Tags_Objects(self):
        if self.filename != '':
            # Define a tabela aberta ativa
            workbookObjects = openpyxl.Workbook()
            sheetObjects = workbookObjects.active
            sheetObjects.title = 'Hmi Tags'
            tag_data = read_column(self.filename, tag_name, 'LISTA DE MEMORIA')
            função_data = read_column(self.filename, função_name, 'LISTA DE MEMORIA')
            tipo_controle_data = read_column(self.filename, tipo_controle, 'LISTA DE MEMORIA')
            fc_db_number_data = read_column(self.filename, fc_db_number, 'LISTA DE MEMORIA')
            area_data = read_column(self.filename, area_controle, 'LISTA DE MEMORIA')
            
            tag_data = read_column(self.filename, tag_name, 'LISTA DE MEMORIA')
            cabecalho = ["Name", "Path", "Connection", "PLC tag", "DataType", "Length", "Coding", "Access Method", "Address", "Indirect addressing", "Index tag", "Start value", "ID tag", "Display name [en-US]", "Comment [en-US]", "Acquisition mode", "Acquisition cycle", "Limit Upper 2 Type", "Limit Upper 2", "Limit Upper 1 Type", "Limit Upper 1", "Limit Lower 1 Type", "Limit Lower 1", "Limit Lower 2 Type","Limit Lower 2", "Linear scaling", "End value PLC", "Start value PLC", "End value HMI", "Start value HMI", "Gmp relevant", "Confirmation Type", "Mandatory Commenting"]
            sheetObjects.append(cabecalho)
            # Write data to the worksheet
            for j in range(len(tag_data)):
                if(tipo_controle_data[j] == "ACM"):
                     sheetObjects.append(['DB' + tag_data[j] + '_Odi_HmiSta', "ACM\ACM_Objects", "HMI_Connection_1", 'DB' + tag_data[j] + '.Odi_HmiSta', "Dint", "4", "Binary", "Symbolic access", "<No value>", "False", "<No value>", "<No value>", "0", "<No value>", "<No value>", "Cyclic in operation", "1s", "None", "<No value>", "None", "<No value>", "False", "10", "0", "100", "0", "False", "None", "False"])
                     sheetObjects.append(['DB' + tag_data[j] + '_Oi_HmiAlm', "ACM\ACM_Objects", "HMI_Connection_1", 'DB' + tag_data[j] + '.Oi_HmiAlm', "Int", "2", "Binary", "Symbolic access", "<No value>", "False", "<No value>", "<No value>", "0", "<No value>", "<No value>", "Continuous", "1s", "None", "<No value>", "None", "<No value>", "False", "10", "0", "100", "0", "False", "None", "False"])
                     sheetObjects.append(['DB' + tag_data[j] + '_Or_SpOut', "ACM\ACM_Objects", "HMI_Connection_1", 'DB' + tag_data[j] + '.Or_SpOut', "Real", "4", "IEEE754", "Symbolic access", "<No value>", "False", "<No value>", "<No value>", "0", "<No value>", "<No value>", "Cyclic in operation", "1s", "None", "<No value>", "None", "<No value>", "False", "10", "0", "100", "0", "False", "None", "False"])
                if(tipo_controle_data[j] == "AV"):
                     sheetObjects.append(['DB' + tag_data[j] + '_Odi_HmiSta', "AV\AV_Objects", "HMI_Connection_1", 'DB' + tag_data[j] + '.Odi_HmiSta', "Dint", "4", "Binary", "Symbolic access", "<No value>", "False", "<No value>", "<No value>", "0", "<No value>", "<No value>", "Cyclic in operation", "1s", "None", "<No value>", "None", "<No value>", "False", "10", "0", "100", "0", "False", "None", "False"])
                     sheetObjects.append(['DB' + tag_data[j] + '_Oi_HmiAlm', "AV\AV_Objects", "HMI_Connection_1", 'DB' + tag_data[j] + '.Oi_HmiAlm', "Int", "2", "Binary", "Symbolic access", "<No value>", "False", "<No value>", "<No value>", "0", "<No value>", "<No value>", "Continuous", "1s", "None", "<No value>", "None", "<No value>", "False", "10", "0", "100", "0", "False", "None", "False"])
                     sheetObjects.append(['DB' + tag_data[j] + '_Or_ProcVal', "AV\AV_Objects", "HMI_Connection_1", 'DB' + tag_data[j] + '.Or_ProcVal', "Real", "4", "IEEE754", "Symbolic access", "<No value>", "False", "<No value>", "<No value>", "0", "<No value>", "<No value>", "Cyclic in operation", "1s", "None", "<No value>", "None", "<No value>", "False", "10", "0", "100", "0", "False", "None", "False"])
                     sheetObjects.append(['DB' + tag_data[j] + '_Ir_ParScMax', "AV\AV_Objects", "HMI_Connection_1", 'DB' + tag_data[j] + '.Ir_ParScMax', "Real", "4", "IEEE754", "Symbolic access", "<No value>", "False", "<No value>", "<No value>", "0", "<No value>", "<No value>", "Cyclic in operation", "1s", "None", "<No value>", "None", "<No value>", "False", "10", "0", "100", "0", "False", "None", "False"])
                     sheetObjects.append(['DB' + tag_data[j] + '_Ir_ParScMin', "AV\AV_Objects", "HMI_Connection_1", 'DB' + tag_data[j] + '.Ir_ParScMin', "Real", "4", "IEEE754", "Symbolic access", "<No value>", "False", "<No value>", "<No value>", "0", "<No value>", "<No value>", "Cyclic in operation", "1s", "None", "<No value>", "None", "<No value>", "False", "10", "0", "100", "0", "False", "None", "False"])
                if(tipo_controle_data[j] == "AXV"):
                     sheetObjects.append(['DB' + tag_data[j] + '_Odi_HmiSta', "AXV\AXV_Objects", "HMI_Connection_1", 'DB' + tag_data[j] + '.Odi_HmiSta', "Dint", "4", "Binary", "Symbolic access", "<No value>", "False", "<No value>", "<No value>", "0", "<No value>", "<No value>", "Cyclic in operation", "1s", "None", "<No value>", "None", "<No value>", "False", "10", "0", "100", "0", "False", "None", "False"])
                     sheetObjects.append(['DB' + tag_data[j] + '_Oi_HmiAlm', "AXV\AXV_Objects", "HMI_Connection_1", 'DB' + tag_data[j] + '.Oi_HmiAlm', "Int", "2", "Binary", "Symbolic access", "<No value>", "False", "<No value>", "<No value>", "0", "<No value>", "<No value>", "Cyclic in operation", "1s", "None", "<No value>", "None", "<No value>", "False", "10", "0", "100", "0", "False", "None", "False"])
                     sheetObjects.append(['DB' + tag_data[j] + '_Ir_ActPos', "AXV\AXV_Objects", "HMI_Connection_1", 'DB' + tag_data[j] + '.Ir_ActPos', "Real", "4", "IEEE754", "Symbolic access", "<No value>", "False", "<No value>", "<No value>", "0", "<No value>", "<No value>", "Cyclic in operation", "1s", "None", "<No value>", "None", "<No value>", "False", "10", "0", "100", "0", "False", "None", "False"])
                if(tipo_controle_data[j] == "DCM"):
                     sheetObjects.append(['DB' + tag_data[j] + '_Odi_HmiSta', "DCM\DCM_Objects", "HMI_Connection_1", 'DB' + tag_data[j] + '.Odi_HmiSta', "Dint", "4", "Binary", "Symbolic access", "<No value>", "False", "<No value>", "<No value>", "0", "<No value>", "<No value>", "Cyclic in operation", "1s", "None", "<No value>", "None", "<No value>", "False", "10", "0", "100", "0", "False", "None", "False"])
                     sheetObjects.append(['DB' + tag_data[j] + '_Oi_HmiAlm', "DCM\DCM_Objects", "HMI_Connection_1", 'DB' + tag_data[j] + '.Oi_HmiAlm', "Int", "2", "Binary", "Symbolic access", "<No value>", "False", "<No value>", "<No value>", "0", "<No value>", "<No value>", "Continuous", "1s", "None", "<No value>", "None", "<No value>", "False", "10", "0", "100", "0", "False", "None", "False"])
                if(tipo_controle_data[j] == "DI"):
                     sheetObjects.append(['DB' + tag_data[j] + '_Odi_HmiSta', "DI\DI_Objects", "HMI_Connection_1", 'DB' + tag_data[j] + '.Odi_HmiSta', "Dint", "4", "Binary", "Symbolic access", "<No value>", "False", "<No value>", "<No value>", "0", "<No value>", "<No value>", "Cyclic in operation", "1s", "None", "<No value>", "None", "<No value>", "False", "10", "0", "100", "0", "False", "None", "False"])
                     sheetObjects.append(['DB' + tag_data[j] + '_Oi_HmiAlm', "DI\DI_Objects", "HMI_Connection_1", 'DB' + tag_data[j] + '.Oi_HmiAlm', "Int", "2", "Binary", "Symbolic access", "<No value>", "False", "<No value>", "<No value>", "0", "<No value>", "<No value>", "Continuous", "1s", "None", "<No value>", "None", "<No value>", "False", "10", "0", "100", "0", "False", "None", "False"])
                if(tipo_controle_data[j] == "DXV"):
                     sheetObjects.append(['DB' + tag_data[j] + '_Odi_HmiSta', "DXV\DXV_Objects", "HMI_Connection_1", 'DB' + tag_data[j] + '.Odi_HmiSta', "Dint", "4", "Binary", "Symbolic access", "<No value>", "False", "<No value>", "<No value>", "0", "<No value>", "<No value>", "Cyclic in operation", "1s", "None", "<No value>", "None", "<No value>", "False", "10", "0", "100", "0", "False", "None", "False"])
                     sheetObjects.append(['DB' + tag_data[j] + '_Oi_HmiAlm', "DXV\DXV_Objects", "HMI_Connection_1", 'DB' + tag_data[j] + '.Oi_HmiAlm', "Int", "2", "Binary", "Symbolic access", "<No value>", "False", "<No value>", "<No value>", "0", "<No value>", "<No value>", "Continuous", "1s", "None", "<No value>", "None", "<No value>", "False", "10", "0", "100", "0", "False", "None", "False"])
                     sheetObjects.append(['DB' + tag_data[j] + '_b_HmiTagVis', "DXV\DXV_Objects", "<No Value>", "<No Value>", "Bool", "1", "Binary", "<No value>", "<No value>", "False", "<No value>", "1", "0", "<No value>", "<No value>", "Cyclic in operation", "1s", "None", "<No value>", "None", "<No value>", "False", "10", "0", "100", "0", "False", "None", "False"])
                     sheetObjects.append(['DB' + tag_data[j] + '_b_Interlock', "DXV\DXV_Objects", "<No Value>", "<No Value>", "Bool", "1", "Binary", "<No value>", "<No value>", "False", "<No value>", "<No value>", "0", "<No value>", "<No value>", "Cyclic in operation", "1s", "None", "<No value>", "None", "<No value>", "False", "10", "0", "100", "0", "False", "None", "False"])
            PATH = os.path.join(current_directory,"GENERATED\\HMI_Objects.xlsx")
            workbookObjects.save(PATH)
            workbookObjects.close()
            self.status_label.setText("Status: Lista 'Objects' Gerada")
            print(PATH)
            print("O arquivo de Excel 'HMI_Objects.xlsx' foi criado com sucesso!!")


    def Alarmes(self):
        if self.filename != '':
            # Define a tabela aberta ativa
            workbookAlarmes = openpyxl.Workbook()
            sheetAlarmes = workbookAlarmes.active
            sheetAlarmes.title = 'DiscreteAlarms'
            tag_data = read_column(self.filename, tag_name, 'LISTA DE MEMORIA')
            função_data = read_column(self.filename, função_name, 'LISTA DE MEMORIA')
            tipo_controle_data = read_column(self.filename, tipo_controle, 'LISTA DE MEMORIA')
            fc_db_number_data = read_column(self.filename, fc_db_number, 'LISTA DE MEMORIA')
            area_data = read_column(self.filename, area_controle, 'LISTA DE MEMORIA')

            # Contador para ID dos alarmes
            c = 1

            # Data to write to the Excel file
            cabecalho = ["ID", "Name", "Alarm text [en-US], Alarm text", "FieldInfo [Alarm text]", "Class", "Trigger tag", "Trigger bit", "Acknowledgement tag", "Acknowledgement bit", "PLC acknowledgement tag", "PLC acknowledgement bit", "Group", "Report", "Info text [en-US]", "Info text"]

            sheetAlarmes.append(cabecalho)

            # Write data to the worksheet
            for j in range(len(tag_data)):
                if(tipo_controle_data[j] == "AV"):
                    sheetAlarmes.append([str(c), tag_data[j]+" - Setpoint_HH", tag_data[j]+" - ALARME MUITO ALTO - "+função_data[j]+" (Usuário logado: <field ref=\"0\" /> )", "\"<ref id = 0; type = AlarmTag; Tag = Current_User_Name; DisplayType = Text; Length = 15;>\n\"", "Falhas_Nv2", '"DB_' + tag_data[j] + '.Oi_HmiAlm"', "0", "<No value>", "0", "<No value>", "0", "<No value>", "True", "<No value>"])
                    c+=1
                    sheetAlarmes.append([str(c), tag_data[j]+" - Setpoint_H", tag_data[j]+" - ALARME ALTO - "+função_data[j]+" (Usuário logado: <field ref=\"0\" /> )", "\"<ref id = 0; type = AlarmTag; Tag = Current_User_Name; DisplayType = Text; Length = 15;>\n\"", "Falhas_Nv1", '"DB_' + tag_data[j] + '.Oi_HmiAlm"', "1", "<No value>", "0", "<No value>", "0", "<No value>", "True", "<No value>"])
                    c+=1
                    sheetAlarmes.append([str(c), tag_data[j]+" - Setpoint_L", tag_data[j]+" - ALARME BAIXO - "+função_data[j]+" (Usuário logado: <field ref=\"0\" /> )", "\"<ref id = 0; type = AlarmTag; Tag = Current_User_Name; DisplayType = Text; Length = 15;>\n\"", "Falhas_Nv1", '"DB_' + tag_data[j] + '.Oi_HmiAlm"', "2", "<No value>", "0", "<No value>", "0", "<No value>", "True", "<No value>"])
                    c+=1
                    sheetAlarmes.append([str(c), tag_data[j]+" - Setpoint_LL", tag_data[j]+" - ALARME MUITO BAIXO - "+função_data[j]+" (Usuário logado: <field ref=\"0\" /> )", "\"<ref id = 0; type = AlarmTag; Tag = Current_User_Name; DisplayType = Text; Length = 15;>\n\"", "Falhas_Nv2", '"DB_' + tag_data[j] + '.Oi_HmiAlm"', "3", "<No value>", "0", "<No value>", "0", "<No value>", "True", "<No value>"])
                    c+=1
                if(tipo_controle_data[j] == "ACM"):
                    sheetAlarmes.append([str(c), tag_data[j]+" - EMG", tag_data[j]+" - EMERGÊNCIA GERAL ACIONADA - "+função_data[j]+" (Usuário logado: <field ref=\"0\" /> )", "\"<ref id = 0; type = AlarmTag; Tag = Current_User_Name; DisplayType = Text; Length = 15;>\n\"", "Falhas_Nv3", '"DB_' + tag_data[j] + '.Oi_HmiAlm"', "3", "<No value>", "0", "<No value>", "0", "<No value>", "True", "<No value>"])
                    c+=1
                    sheetAlarmes.append([str(c), tag_data[j]+" - FeedBack", tag_data[j]+" - FEEDBACK NÂO RECEBIDO - "+função_data[j]+" (Usuário logado: <field ref=\"0\" /> )", "\"<ref id = 0; type = AlarmTag; Tag = Current_User_Name; DisplayType = Text; Length = 15;>\n\"", "Falhas_Nv1", '"DB_' + tag_data[j] + '.Oi_HmiAlm"', "0", "<No value>", "0", "<No value>", "0", "<No value>", "True", "<No value>"])
                    c+=1
                    sheetAlarmes.append([str(c), tag_data[j]+" - InvFalha", tag_data[j]+" - FALHA NA BOMBA - "+função_data[j]+" (Usuário logado: <field ref=\"0\" /> )", "\"<ref id = 0; type = AlarmTag; Tag = Current_User_Name; DisplayType = Text; Length = 15;>\n\"", "Falhas_Nv2", '"DB_' + tag_data[j] + '.Oi_HmiAlm"', "4", "<No value>", "0", "<No value>", "0", "<No value>", "True", "<No value>"])
                    c+=1
                if(tipo_controle_data[j] == "DXV"):
                    sheetAlarmes.append([str(c), tag_data[j]+" - Falha_Abrir", tag_data[j]+" - FALHA FEEDBACK - "+função_data[j]+" (Usuário logado: <field ref=\"0\" /> )", "\"<ref id = 0; type = AlarmTag; Tag = Current_User_Name; DisplayType = Text; Length = 15;>\n\"", "Falhas_Nv2", '"DB_' + tag_data[j] + '.Oi_HmiAlm"', "0", "<No value>", "0", "<No value>", "0", "<No value>", "True", "<No value>"])
                    c+=1
                    sheetAlarmes.append([str(c), tag_data[j]+" - Falha_Fechar", tag_data[j]+" - FALHA INCONSISTÊNCIA SENSORES - "+função_data[j]+" (Usuário logado: <field ref=\"0\" /> )", "\"<ref id = 0; type = AlarmTag; Tag = Current_User_Name; DisplayType = Text; Length = 15;>\n\"", "Falhas_Nv2", '"DB_' + tag_data[j] + '.Oi_HmiAlm"', "1", "<No value>", "0", "<No value>", "0", "<No value>", "True", "<No value>"])
                    c+=1
                if(tipo_controle_data[j] == "AXV"):
                    sheetAlarmes.append([str(c), tag_data[j]+" - Feedback_Timeout", tag_data[j]+" - TEMPO DE FEEDBACK EXPIRADO (ABERTURA DA VÁLVULA MUITO LENTA OU PROBLEMAS NOS LIMITES FÍSICOS) - "+função_data[j]+" (Usuário logado: <field ref=\"0\" /> )", "\"<ref id = 0; type = AlarmTag; Tag = Current_User_Name; DisplayType = Text; Length = 15;>\n\"", "Falhas_Nv1", '"DB_' + tag_data[j] + '.Oi_HmiAlm"', "0", "<No value>", "0", "<No value>", "0", "<No value>", "True", "<No value>"])
                    c+=1
                    sheetAlarmes.append([str(c), tag_data[j]+" - Feedback_Inconsistente", tag_data[j]+" - LIMITE DE FEEDBACK DIGITAL NÃO CONSISTENTE - "+função_data[j]+" (Usuário logado: <field ref=\"0\" /> )", "\"<ref id = 0; type = AlarmTag; Tag = Current_User_Name; DisplayType = Text; Length = 15;>\n\"", "Falhas_Nv1", '"DB_' + tag_data[j] + '.Oi_HmiAlm"', "1", "<No value>", "0", "<No value>", "0", "<No value>", "True", "<No value>"])
                    c+=1
                    sheetAlarmes.append([str(c), tag_data[j]+" - Intertravada", tag_data[j]+" - VÁLVULA INTERTRAVADA PARA ABERTURA OU FECHAMENTO - "+função_data[j]+" (Usuário logado: <field ref=\"0\" /> )", "\"<ref id = 0; type = AlarmTag; Tag = Current_User_Name; DisplayType = Text; Length = 15;>\n\"", "Falhas_Nv1", '"DB_' + tag_data[j] + '.Oi_HmiAlm"', "2", "<No value>", "0", "<No value>", "0", "<No value>", "True", "<No value>"])
                    c+=1
                    sheetAlarmes.append([str(c), tag_data[j]+" - Setpoint", tag_data[j]+" - O SETPOINT REQUERIDO PARA ABERTURA E FECHAMENTO ESTÁ FORA DOS LIMITES DE SEGURANÇA - "+função_data[j]+" (Usuário logado: <field ref=\"0\" /> )", "\"<ref id = 0; type = AlarmTag; Tag = Current_User_Name; DisplayType = Text; Length = 15;>\n\"", "Falhas_Nv1", '"DB_' + tag_data[j] + '.Oi_HmiAlm"', "3", "<No value>", "0", "<No value>", "0", "<No value>", "True", "<No value>"])
                    c+=1
                    sheetAlarmes.append([str(c), tag_data[j]+" - Posicionador_analógico", tag_data[j]+" - O POSICIONAMENTO ANALÓGICO ESTÁ FORA DOS LIMITES E INCONSISTENTE COM OS LIMITES DIGITAIS - "+função_data[j]+" (Usuário logado: <field ref=\"0\" /> )", "\"<ref id = 0; type = AlarmTag; Tag = Current_User_Name; DisplayType = Text; Length = 15;>\n\"", "Falhas_Nv1", '"DB_' + tag_data[j] + '.Oi_HmiAlm"', "4", "<No value>", "0", "<No value>", "0", "<No value>", "True", "<No value>"])
                    c+=1
                    sheetAlarmes.append([str(c), tag_data[j]+" - Posição", tag_data[j]+" - A POSIÇÃO ATUAL DA VÁLVULA ESTÁ FORA DOS LIMITES CONFIGURADOS - "+função_data[j]+" (Usuário logado: <field ref=\"0\" /> )", "\"<ref id = 0; type = AlarmTag; Tag = Current_User_Name; DisplayType = Text; Length = 15;>\n\"", "Falhas_Nv1", '"DB_' + tag_data[j] + '.Oi_HmiAlm"', "5", "<No value>", "0", "<No value>", "0", "<No value>", "True", "<No value>"])
                    c+=1
                if(tipo_controle_data[j] == "DI"):
                    sheetAlarmes.append([str(c), tag_data[j]+" - Alarme", tag_data[j]+" - ALARME DE SET POINT ATINGIDO - "+função_data[j]+" (Usuário logado: <field ref=\"0\" /> )", "\"<ref id = 0; type = AlarmTag; Tag = Current_User_Name; DisplayType = Text; Length = 15;>\n\"", "Falhas_Nv3", '"DB_' + tag_data[j] + '.Oi_HmiAlm"', "0", "<No value>", "0", "<No value>", "0", "<No value>", "True", "<No value>"])
                    c+=1
                    sheetAlarmes.append([str(c), tag_data[j]+" - Aviso", tag_data[j]+" - AVISO DE SET POINT ATINGIDO - "+função_data[j]+" (Usuário logado: <field ref=\"0\" /> )", "\"<ref id = 0; type = AlarmTag; Tag = Current_User_Name; DisplayType = Text; Length = 15;>\n\"", "Falhas_Nv2", '"DB_' + tag_data[j] + '.Oi_HmiAlm"', "1", "<No value>", "0", "<No value>", "0", "<No value>", "True", "<No value>"])
                    c+=1
                if(tipo_controle_data[j] == "DCM"):
                    sheetAlarmes.append([str(c), tag_data[j]+" - FeedBack", tag_data[j]+" - FALHA FEEDBACK - "+função_data[j]+" (Usuário logado: <field ref=\"0\" /> )", "\"<ref id = 0; type = AlarmTag; Tag = Current_User_Name; DisplayType = Text; Length = 15;>\n\"", "Falhas_Nv2", '"DB_' + tag_data[j] + '.Oi_HmiAlm"', "0", "<No value>", "0", "<No value>", "0", "<No value>", "True", "<No value>"])
                    c+=1
                    sheetAlarmes.append([str(c), tag_data[j]+" - Falha_Drive", tag_data[j]+" - ERRO DE DRIVE - "+função_data[j]+" (Usuário logado: <field ref=\"0\" /> )", "\"<ref id = 0; type = AlarmTag; Tag = Current_User_Name; DisplayType = Text; Length = 15;>\n\"", "Falhas_Nv2", '"DB_' + tag_data[j] + '.Oi_HmiAlm"', "6", "<No value>", "0", "<No value>", "0", "<No value>", "True", "<No value>"])
                    c+=1
            PATH = os.path.join(current_directory,"GENERATED\\Alarmes IHM.xlsx")
            workbookAlarmes.save(PATH)
            workbookAlarmes.close()
            self.status_label.setText("Status: Alarmes gerados")
            print(PATH)
            print("O arquivo de Excel 'Alarmes IHM.xlsx' foi criado com sucesso!!")

    def Tags(self):
        if self.filename != '':
            # Chamada da função de leitura de do arquivo de excel
            tag_data = read_column(self.filename, tag_name, 'LISTA DE IO')
            tag_data2 = read_column(self.filename, tag_name, 'TAGS PADRAO')
            tag_data.extend(tag_data2)

            função_data = read_column(self.filename, função_name, 'LISTA DE IO')
            função_data2 = read_column(self.filename, função_name, 'TAGS PADRAO')
            função_data.extend(função_data2)

            tipo_controle_data = read_column(self.filename, tipo_controle_tag, 'LISTA DE IO')
            tipo_controle_data2 = read_column(self.filename, tipo_controle_tag, 'TAGS PADRAO')
            tipo_controle_data.extend(tipo_controle_data2)

            ENDEREÇO_number_data = read_column(self.filename, ENDEREÇO_number, 'LISTA DE IO')
            ENDEREÇO_number_data2 = read_column(self.filename, ENDEREÇO_number, 'TAGS PADRAO')
            ENDEREÇO_number_data.extend(ENDEREÇO_number_data2)
         
            # Cria um novo diretório do excel
            workbooktag = openpyxl.Workbook()
            if 'PLC Tags' in workbooktag.sheetnames:
                workbooktag.remove('PLC Tags')
            workbooktag.create_sheet('PLC Tags')

            sheet = workbooktag.active
            sheet = workbooktag["PLC Tags"]
            sheet.title = 'PLC Tags'


            # Data to write to the Excel file
            cabecalho = ["Name", "Path", "Data Type", "Logical Address", "Comment", "Hmi Visible", "Hmi Accessible", "Hmi Writeable", "Typeobject ID", "Version ID"]
            sheet.append(cabecalho)
            # Write data to the worksheet
            for j in range(len(tag_data)):
                if(tipo_controle_data[j] == "DI"):
                    sheet.append([tag_data[j], "Entradas Digitais", "Bool",ENDEREÇO_number_data[j], função_data[j],"True", "True", "True", '',''])
                if(tipo_controle_data[j] == "DO"):
                    sheet.append([tag_data[j], "Saidas Digitais", "Bool",ENDEREÇO_number_data[j], função_data[j],"True", "True", "True", '',''])
                if(tipo_controle_data[j] == "AI"):
                    sheet.append([tag_data[j], "Entradas Analógicas", "Word",ENDEREÇO_number_data[j], função_data[j],"True", "True", "True", '',''])
                if(tipo_controle_data[j] == "AO"):
                    sheet.append([tag_data[j], "Saidas Analógicas", "Word",ENDEREÇO_number_data[j], função_data[j],"True", "True", "True", '',''])
                if(tipo_controle_data[j] == "PD"):
                    sheet.append([tag_data[j], "Tags Padrão", "Bool",ENDEREÇO_number_data[j], função_data[j],"True", "True", "True", '',''])

            if 'TagTable Properties' in workbooktag.sheetnames:
                workbooktag.remove('TagTable Properties')
            workbooktag.create_sheet("TagTable Properties")
            sheet1 = workbooktag["TagTable Properties"]

            cabecalho1 = ["Path", "BelongsToUnit", "Accessibility"]
            sheet1.append(cabecalho1)
            sheet1.append(["Entradas Digitais","",""])
            sheet1.append(["Saidas Digitais","",""])
            sheet1.append(["Entradas Analógicas","",""])
            sheet1.append(["Saidas Analógicas","",""])
            sheet1.append(["Tags Padrão","",""])

            # Save the workbook
            PATHtag = os.path.join(current_directory,"GENERATED\\Lista de TAGs PLC.xlsx")
            workbooktag.save(PATHtag)
            workbooktag.close()
            self.status_label.setText("Status: Tags gerados")
            print("O arquivo de Excel 'LISTA DE TAGs PLC.xlsx' foi criado com sucesso!!")

    def HMI_Tags(self):
            
            Alarmes_Path = os.path.join(current_directory,"GENERATED\\Alarmes IHM.xlsx")
            pasta_origem = os.path.join(current_directory, "SOURCE")
            pasta_destino = os.path.join(current_directory, "GENERATED")
            
            #copia a pasta "PopUp_Tags" da source para a generated
            for nome_arquivo in os.listdir(pasta_origem):
                if nome_arquivo.endswith(".xlsx"):
                    caminho_origem = os.path.join(pasta_origem, nome_arquivo)
                    caminho_destino = os.path.join(pasta_destino, nome_arquivo)

                    shutil.copy2(caminho_origem, caminho_destino)

            if self.filename != '':
                # Chamada da função de leitura de do arquivo de excel
                tag_data = read_column(self.filename, tag_name, 'HMI TAGS')
                tag_plc = read_column(self.filename,'PLC tag', 'HMI TAGS')
                data_type = read_column(self.filename,'Data Type', 'HMI TAGS')
                data_length = read_column(self.filename,'Length', 'HMI TAGS')

                tag_alarm_data_raw = read_column(Alarmes_Path,'Trigger tag','DiscreteAlarms')
                tag_alarm_data = [line.replace('"','') for line in tag_alarm_data_raw]
                tag_alarm_plc = tag_alarm_data
            
                # Cria um novo diretório do excel
                workbooktag = openpyxl.Workbook()
                if 'Hmi Tags' in workbooktag.sheetnames:
                    workbooktag.remove('Hmi Tags')
                workbooktag.create_sheet('Hmi Tags')

                sheet = workbooktag.active
                sheet = workbooktag["Hmi Tags"]
                sheet.title = 'Hmi Tags'

            
                # Data to write to the Excel file
            cabecalho = ["Name", "Path", "Connection", "PLC tag", "DataType", "Length", "Coding", "Access Method", "Address", "Indirect addressing","Index tag","Start value","ID tag","Display name [en-US]","Comment [en-US]","Acquisition mode","Acquisition cycle","Limit Upper 2 Type","Limit Upper 2","Limit Upper 1 Type","Limit Upper 1","Limit Lower 1 Type","Limit Lower 1","Limit Lower 2 Type","Limit Lower 2","Linear scaling","End value PLC","Start value PLC","End value HMI","Start value HMI","Gmp relevant","Confirmation Type","Mandatory Commenting"]
            sheet.append(cabecalho)
            # Write data to the worksheet
            for j in range(len(tag_data)):
                sheet.append([tag_data[j], 'Default tag table', '<No Value>',tag_plc[j],data_type[j],data_length[j],'Binary','<No Value>','<No Value>','False','<No Value>','<No Value>','0','<No Value>','<No Value>','Cyclic in operation','100 ms','None','<No Value>','None','<No Value>','None','<No Value>','None','<No Value>','False','10','0','100','False','None','False'])
            if 'Multiplexing' in workbooktag.sheetnames:
                workbooktag.remove('Multiplexing')
            workbooktag.create_sheet("Multiplexing")
            sheet1 = workbooktag["Multiplexing"]

            cabecalho1 = ["HMI Tag name", "Multiplex Tag", "Index"]
            sheet1.append(cabecalho1)

            # Save the workbook
            PATHtag = os.path.join(current_directory,"GENERATED\\Lista de TAGs HMI.xlsx")
            workbooktag.save(PATHtag)
            workbooktag.close()
            self.status_label.setText("Status: Hmi Tags gerados")
            print("O arquivo de Excel 'LISTA DE TAGs HMI.xlsx' foi criado com sucesso!!")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    win = MainWin()
    sys.exit(app.exec_())